#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar reportes de totalidad por sociedad.
Procesa archivos CSV de libro diario y sumas y saldos, generando un Excel
con validación de cuadres contables.
"""

import pandas as pd
import os
from pathlib import Path
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import warnings

warnings.filterwarnings('ignore')


class GeneradorTotalidad:
    """Clase para generar reportes de totalidad por sociedad."""

    def __init__(self, ruta_datos_tratados: str, ruta_salida: str):
        """
        Inicializa el generador de totalidad.

        Args:
            ruta_datos_tratados: Ruta a la carpeta con datos procesados
            ruta_salida: Ruta donde se guardarán los reportes de totalidad
        """
        self.ruta_datos_tratados = Path(ruta_datos_tratados)
        self.ruta_salida = Path(ruta_salida)

        # Crear carpeta de salida si no existe
        self.ruta_salida.mkdir(parents=True, exist_ok=True)

    def buscar_archivos_por_sociedad(self) -> Dict[str, Dict[str, List[Path]]]:
        """
        Busca y agrupa archivos CSV por sociedad.

        Returns:
            Diccionario con estructura: {nombre_sociedad: {'libro_diario': [paths], 'sumas_saldos': [paths]}}
        """
        sociedades = {}

        # Recorrer todas las carpetas en datos_tratados
        for carpeta_sociedad in self.ruta_datos_tratados.iterdir():
            if not carpeta_sociedad.is_dir():
                continue

            nombre_sociedad = carpeta_sociedad.name

            # Buscar archivos de libro diario y sumas y saldos
            archivos_ld = list(carpeta_sociedad.glob('**/libro_diario_*.csv'))
            archivos_sys = list(carpeta_sociedad.glob('**/sumas_saldos_*.csv'))

            if archivos_ld or archivos_sys:
                sociedades[nombre_sociedad] = {
                    'libro_diario': archivos_ld,
                    'sumas_saldos': archivos_sys
                }

        return sociedades

    def convertir_a_numerico(self, df: pd.DataFrame, columnas: List[str]) -> pd.DataFrame:
        """
        Convierte las columnas especificadas a tipo numérico.

        Args:
            df: DataFrame a procesar
            columnas: Lista de nombres de columnas a convertir

        Returns:
            DataFrame con columnas convertidas
        """
        df_resultado = df.copy()

        for columna in columnas:
            if columna in df_resultado.columns:
                df_resultado[columna] = pd.to_numeric(df_resultado[columna], errors='coerce').fillna(0)

        return df_resultado

    def procesar_libro_diario(self, archivos_ld: List[Path]) -> pd.DataFrame:
        """
        Procesa uno o más archivos de libro diario y crea las columnas GT_.

        Args:
            archivos_ld: Lista de paths a archivos CSV de libro diario

        Returns:
            DataFrame procesado con columnas GT_
        """
        # Consolidar todos los archivos de libro diario
        df_list = []
        for archivo in archivos_ld:
            try:
                df = pd.read_csv(archivo)
                df_list.append(df)
            except Exception as e:
                print(f"⚠️  Error leyendo {archivo.name}: {e}")

        if not df_list:
            return pd.DataFrame()

        df_diario = pd.concat(df_list, ignore_index=True)

        # Mapear columnas a formato GT_
        # Buscar columnas de debe, haber, cuenta y asiento
        col_debe = None
        col_haber = None
        col_cuenta = None
        col_asiento = None

        for col in df_diario.columns:
            col_lower = col.lower()
            if 'debe' in col_lower and 'moneda local' in col_lower and col_debe is None:
                col_debe = col
            elif 'haber' in col_lower and 'moneda local' in col_lower and col_haber is None:
                col_haber = col
            elif col_lower == 'cuenta' and col_cuenta is None:
                col_cuenta = col
            elif col_lower in ['número', 'numero', 'asiento'] and col_asiento is None:
                col_asiento = col

        # Crear columnas GT_
        df_diario['GT_DEBE'] = pd.to_numeric(df_diario[col_debe] if col_debe else 0, errors='coerce').fillna(0)
        df_diario['GT_HABER'] = pd.to_numeric(df_diario[col_haber] if col_haber else 0, errors='coerce').fillna(0)
        df_diario['GT_IMPORTE_MONEDA_LOCAL'] = df_diario['GT_DEBE'] - df_diario['GT_HABER']
        df_diario['GT_CUENTA'] = df_diario[col_cuenta].astype(str) if col_cuenta else 'Sin_Cuenta'
        df_diario['GT_ASIENTO'] = df_diario[col_asiento].astype(str) if col_asiento else 'Sin_Asiento'

        return df_diario

    def procesar_sumas_saldos(self, archivos_sys: List[Path]) -> pd.DataFrame:
        """
        Procesa uno o más archivos de sumas y saldos y crea las columnas GT_.

        Args:
            archivos_sys: Lista de paths a archivos CSV de sumas y saldos

        Returns:
            DataFrame procesado con columnas GT_
        """
        # Consolidar todos los archivos de sumas y saldos
        df_list = []
        for archivo in archivos_sys:
            try:
                df = pd.read_csv(archivo)
                df_list.append(df)
            except Exception as e:
                print(f"⚠️  Error leyendo {archivo.name}: {e}")

        if not df_list:
            return pd.DataFrame()

        df_sumas = pd.concat(df_list, ignore_index=True)

        # Mapear columnas a formato GT_
        col_cuenta = None
        col_arrastre = None
        col_periodos_ant = None
        col_debe_sys = None
        col_haber_sys = None
        col_saldo_periodo = None

        for col in df_sumas.columns:
            col_lower = col.lower()
            if 'cta' in col_lower and 'mayor' in col_lower and col_cuenta is None:
                col_cuenta = col
            elif 'arrastre' in col_lower and 'saldo' in col_lower and col_arrastre is None:
                col_arrastre = col
            elif ('saldo' in col_lower or 'per') and 'anterior' in col_lower and col_periodos_ant is None:
                col_periodos_ant = col
            elif 'debe' in col_lower and ('período' in col_lower or 'periodo' in col_lower or 'per.inf' in col_lower) and col_debe_sys is None:
                col_debe_sys = col
            elif 'haber' in col_lower and ('período' in col_lower or 'periodo' in col_lower or 'per.inf' in col_lower) and col_haber_sys is None:
                col_haber_sys = col
            elif 'saldo acumulado' in col_lower and col_saldo_periodo is None:
                col_saldo_periodo = col

        # Crear columnas GT_
        df_sumas['GT_CUENTA'] = df_sumas[col_cuenta].astype(str) if col_cuenta else 'Sin_Cuenta'
        df_sumas['GT_ARRASTRE_SALDOS'] = pd.to_numeric(df_sumas[col_arrastre] if col_arrastre else 0, errors='coerce').fillna(0)
        df_sumas['GT_PERIODOS_ANTERIORES'] = pd.to_numeric(df_sumas[col_periodos_ant] if col_periodos_ant else 0, errors='coerce').fillna(0)
        df_sumas['GT_SALDO_DEBE_SyS'] = pd.to_numeric(df_sumas[col_debe_sys] if col_debe_sys else 0, errors='coerce').fillna(0)
        df_sumas['GT_SALDO_HABER_SyS'] = pd.to_numeric(df_sumas[col_haber_sys] if col_haber_sys else 0, errors='coerce').fillna(0)
        df_sumas['GT_SALDO_PERIODO_SyS'] = pd.to_numeric(df_sumas[col_saldo_periodo] if col_saldo_periodo else 0, errors='coerce').fillna(0)

        return df_sumas

    def aplicar_formato_tabla(self, worksheet, dataframe, rango_inicio: str, nombre_tabla: str):
        """
        Aplica formato de tabla a un rango de celdas.

        Args:
            worksheet: Hoja de Excel
            dataframe: DataFrame con los datos
            rango_inicio: Celda de inicio (ej: 'A1')
            nombre_tabla: Nombre único para la tabla
        """
        if dataframe.empty:
            return

        num_filas, num_columnas = dataframe.shape

        # Convertir rango_inicio a columna y fila
        col_letra = ''.join(filter(str.isalpha, rango_inicio))
        fila_inicio = int(''.join(filter(str.isdigit, rango_inicio)))

        # Calcular columna final
        col_inicio_idx = ord(col_letra) - ord('A')
        col_fin_idx = col_inicio_idx + num_columnas - 1

        # Manejar columnas más allá de Z
        if col_fin_idx <= 25:
            col_fin = chr(col_fin_idx + ord('A'))
        else:
            # Para columnas AA, AB, etc.
            primera_letra = chr((col_fin_idx // 26) - 1 + ord('A'))
            segunda_letra = chr((col_fin_idx % 26) + ord('A'))
            col_fin = primera_letra + segunda_letra

        # El rango incluye encabezados
        rango = f"{col_letra}{fila_inicio}:{col_fin}{fila_inicio + num_filas}"

        # Crear tabla
        tabla = Table(displayName=nombre_tabla, ref=rango)
        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tabla.tableStyleInfo = estilo
        worksheet.add_table(tabla)

        # Aplicar formato numérico a columnas numéricas
        for row_idx in range(fila_inicio + 1, fila_inicio + num_filas + 1):
            for col_idx in range(num_columnas):
                celda = worksheet.cell(row=row_idx, column=col_inicio_idx + col_idx + 1)
                # Si el valor es numérico, aplicar formato
                if isinstance(celda.value, (int, float)):
                    celda.number_format = '#,##0.00'

    def generar_excel_totalidad(self, nombre_sociedad: str, df_diario: pd.DataFrame,
                                df_sumas: pd.DataFrame) -> Tuple[bool, str]:
        """
        Genera el archivo Excel de totalidad para una sociedad.

        Args:
            nombre_sociedad: Nombre de la sociedad
            df_diario: DataFrame del libro diario procesado
            df_sumas: DataFrame de sumas y saldos procesado

        Returns:
            Tupla (validacion_exitosa, ruta_archivo)
        """
        archivo_salida = self.ruta_salida / f"Totalidad_{nombre_sociedad}.xlsx"

        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto

        # HOJA 1: Resumen Total
        ws1 = wb.create_sheet("Resumen_Total")
        total_importe = df_diario['GT_IMPORTE_MONEDA_LOCAL'].sum()

        df_resumen_total = pd.DataFrame({
            'GT_IMPORTE_MONEDA_LOCAL': [total_importe]
        })

        for r_idx, row in enumerate(dataframe_to_rows(df_resumen_total, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                celda = ws1.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Encabezado
                    celda.font = Font(bold=True)
                else:
                    celda.number_format = '#,##0.00'

        ws1.column_dimensions['A'].width = 30
        self.aplicar_formato_tabla(ws1, df_resumen_total, 'A1', f'Tabla_ResumenTotal_{nombre_sociedad.replace(" ", "_")}')

        # HOJA 2: Resumen por Asiento
        ws2 = wb.create_sheet("Resumen_Por_Asiento")
        resumen_asiento = df_diario.groupby('GT_ASIENTO').agg({
            'GT_DEBE': 'sum',
            'GT_HABER': 'sum',
            'GT_IMPORTE_MONEDA_LOCAL': 'sum'
        }).reset_index().round(2)

        for r_idx, row in enumerate(dataframe_to_rows(resumen_asiento, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                celda = ws2.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Encabezado
                    celda.font = Font(bold=True)
                elif c_idx > 1:  # Columnas numéricas
                    celda.number_format = '#,##0.00'

        ws2.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D']:
            ws2.column_dimensions[col].width = 18
        self.aplicar_formato_tabla(ws2, resumen_asiento, 'A1', f'Tabla_ResumenAsiento_{nombre_sociedad.replace(" ", "_")}')

        # HOJA 3: Resumen por Cuenta
        ws3 = wb.create_sheet("Resumen_Por_Cuenta")

        # Resumen del libro diario por cuenta
        resumen_cuenta = df_diario.groupby('GT_CUENTA').agg({
            'GT_DEBE': 'sum',
            'GT_HABER': 'sum',
            'GT_IMPORTE_MONEDA_LOCAL': 'sum'
        }).reset_index()

        # Resumen de sumas y saldos por cuenta
        resumen_sumas_cuenta = df_sumas.groupby('GT_CUENTA').agg({
            'GT_PERIODOS_ANTERIORES': 'sum',
            'GT_ARRASTRE_SALDOS': 'sum',
            'GT_SALDO_DEBE_SyS': 'sum',
            'GT_SALDO_HABER_SyS': 'sum',
            'GT_SALDO_PERIODO_SyS': 'sum'
        }).reset_index()

        # JOIN
        resumen_final = pd.merge(
            resumen_cuenta,
            resumen_sumas_cuenta,
            on='GT_CUENTA',
            how='outer'
        ).fillna(0).round(2)

        # Calcular diferencia
        resumen_final['GT_DIFERENCIA'] = (
            resumen_final['GT_IMPORTE_MONEDA_LOCAL'] +
            resumen_final['GT_ARRASTRE_SALDOS']
        ) - resumen_final['GT_SALDO_PERIODO_SyS']
        resumen_final['GT_DIFERENCIA'] = resumen_final['GT_DIFERENCIA'].round(2)

        for r_idx, row in enumerate(dataframe_to_rows(resumen_final, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                celda = ws3.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Encabezado
                    celda.font = Font(bold=True)
                elif c_idx > 1:  # Columnas numéricas
                    celda.number_format = '#,##0.00'

        ws3.column_dimensions['A'].width = 20
        for col_letra in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws3.column_dimensions[col_letra].width = 20
        self.aplicar_formato_tabla(ws3, resumen_final, 'A1', f'Tabla_ResumenCuenta_{nombre_sociedad.replace(" ", "_")}')

        # Verificar validación (diferencias cercanas a cero)
        no_cumplen = sum(abs(resumen_final['GT_DIFERENCIA']) >= 0.01)
        validacion_exitosa = no_cumplen <= 2

        # HOJA 4: Documentación
        ws4 = wb.create_sheet("Documentacion")

        doc_data = [
            ['DOCUMENTACIÓN DEL REPORTE DE TOTALIDAD', ''],
            ['', ''],
            ['Sociedad:', nombre_sociedad],
            ['Fecha de generación:', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['CÁLCULOS REALIZADOS', ''],
            ['', ''],
            ['Hoja 1 - Resumen Total:', ''],
            ['  GT_IMPORTE_MONEDA_LOCAL', '= SUMA(GT_DEBE - GT_HABER) de todo el libro diario'],
            ['  Debe ser:', '0,00 (o muy cercano a 0)'],
            ['', ''],
            ['Hoja 2 - Resumen por Asiento:', ''],
            ['  GT_DEBE', '= Suma del debe por asiento'],
            ['  GT_HABER', '= Suma del haber por asiento'],
            ['  GT_IMPORTE_MONEDA_LOCAL', '= GT_DEBE - GT_HABER (debe ser 0,00 por asiento)'],
            ['', ''],
            ['Hoja 3 - Resumen por Cuenta:', ''],
            ['  GT_DEBE', '= Suma del debe por cuenta (del libro diario)'],
            ['  GT_HABER', '= Suma del haber por cuenta (del libro diario)'],
            ['  GT_IMPORTE_MONEDA_LOCAL', '= GT_DEBE - GT_HABER'],
            ['  GT_PERIODOS_ANTERIORES', '= Del archivo sumas y saldos'],
            ['  GT_ARRASTRE_SALDOS', '= Del archivo sumas y saldos'],
            ['  GT_SALDO_DEBE_SyS', '= Del archivo sumas y saldos'],
            ['  GT_SALDO_HABER_SyS', '= Del archivo sumas y saldos'],
            ['  GT_SALDO_PERIODO_SyS', '= Del archivo sumas y saldos'],
            ['  GT_DIFERENCIA', '= (GT_IMPORTE_MONEDA_LOCAL + GT_ARRASTRE_SALDOS) - GT_SALDO_PERIODO_SyS'],
            ['  ', 'Debe ser 0,00 o muy cercano a 0'],
            ['', ''],
            ['VALIDACIÓN:', 'EXITOSA' if validacion_exitosa else 'NO EXITOSA'],
            ['Diferencias encontradas:', str(no_cumplen)],
            ['Tolerancia:', 'Máximo 2 diferencias >= 0.01'],
        ]

        for r_idx, row_data in enumerate(doc_data, 1):
            for c_idx, value in enumerate(row_data, 1):
                celda = ws4.cell(row=r_idx, column=c_idx, value=value)
                # Formato especial para títulos
                if 'DOCUMENTACIÓN' in str(value) or 'CÁLCULOS' in str(value):
                    celda.font = Font(bold=True, size=14)
                elif value.endswith(':') and c_idx == 1:
                    celda.font = Font(bold=True)

        ws4.column_dimensions['A'].width = 35
        ws4.column_dimensions['B'].width = 60

        # Guardar archivo
        wb.save(archivo_salida)

        simbolo = "✅" if validacion_exitosa else "❌"
        print(f"{simbolo} Totalidad generada: {nombre_sociedad} - Validación: {'EXITOSA' if validacion_exitosa else 'NO EXITOSA'}")

        return validacion_exitosa, str(archivo_salida)

    def procesar_todas_las_sociedades(self):
        """Procesa todas las sociedades encontradas y genera reportes de totalidad."""
        print("\n" + "="*70)
        print("🚀 INICIANDO GENERACIÓN DE TOTALIDAD")
        print("="*70)

        sociedades = self.buscar_archivos_por_sociedad()

        if not sociedades:
            print("⚠️  No se encontraron sociedades para procesar")
            return

        print(f"📊 Se encontraron {len(sociedades)} sociedades para procesar\n")

        resultados = {
            'exitosas': [],
            'no_exitosas': [],
            'errores': []
        }

        for nombre_sociedad, archivos in sociedades.items():
            try:
                print(f"Procesando: {nombre_sociedad}")

                archivos_ld = archivos.get('libro_diario', [])
                archivos_sys = archivos.get('sumas_saldos', [])

                if not archivos_ld:
                    print(f"⚠️  No se encontró libro diario para {nombre_sociedad}")
                    resultados['errores'].append(nombre_sociedad)
                    continue

                if not archivos_sys:
                    print(f"⚠️  No se encontró sumas y saldos para {nombre_sociedad}")
                    resultados['errores'].append(nombre_sociedad)
                    continue

                # Procesar archivos
                df_diario = self.procesar_libro_diario(archivos_ld)
                df_sumas = self.procesar_sumas_saldos(archivos_sys)

                if df_diario.empty:
                    print(f"⚠️  Libro diario vacío para {nombre_sociedad}")
                    resultados['errores'].append(nombre_sociedad)
                    continue

                # Generar Excel
                validacion_exitosa, ruta_archivo = self.generar_excel_totalidad(
                    nombre_sociedad, df_diario, df_sumas
                )

                if validacion_exitosa:
                    resultados['exitosas'].append(nombre_sociedad)
                else:
                    resultados['no_exitosas'].append(nombre_sociedad)

            except Exception as e:
                print(f"❌ Error procesando {nombre_sociedad}: {e}")
                import traceback
                traceback.print_exc()
                resultados['errores'].append(nombre_sociedad)

        # Resumen final
        print("\n" + "="*70)
        print("📈 RESUMEN DE TOTALIDAD")
        print("="*70)
        print(f"✅ Validaciones exitosas: {len(resultados['exitosas'])}")
        print(f"❌ Validaciones no exitosas: {len(resultados['no_exitosas'])}")
        print(f"⚠️  Errores: {len(resultados['errores'])}")
        print("="*70)

        if resultados['exitosas']:
            print("\n✅ Sociedades con validación EXITOSA:")
            for sociedad in resultados['exitosas']:
                print(f"   - {sociedad}")

        if resultados['no_exitosas']:
            print("\n❌ Sociedades con validación NO EXITOSA:")
            for sociedad in resultados['no_exitosas']:
                print(f"   - {sociedad}")

        if resultados['errores']:
            print("\n⚠️  Sociedades con ERRORES:")
            for sociedad in resultados['errores']:
                print(f"   - {sociedad}")


def main():
    """Función principal."""
    generador = GeneradorTotalidad(
        ruta_datos_tratados='datos_tratados',
        ruta_salida='totalidad'
    )

    generador.procesar_todas_las_sociedades()


if __name__ == '__main__':
    main()
